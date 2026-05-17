import streamlit as st
import pdfplumber
import re
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import datetime

st.set_page_config(page_title="คำนวณโหลดสินค้า IFO", page_icon="📦", layout="wide")

# ── CONSTANTS ──
BOX_CANVAS = 12
SACK_200 = 120
BOX_212 = 24

MONTH_MAP = {
    'ม.ค.':'01','ก.พ.':'02','มี.ค.':'03','เม.ย.':'04',
    'พ.ค.':'05','มิ.ย.':'06','ก.ค.':'07','ส.ค.':'08',
    'ก.ย.':'09','ต.ค.':'10','พ.ย.':'11','ธ.ค.':'12'
}

def clean(s):
    s = re.sub(r'\(cid:\d+\)', '', s)
    return re.sub(r'\s+', ' ', s).strip()

def th_date(s):
    if not s: return ''
    try:
        y,m,d = s.split('-')
        mn=['','ม.ค.','ก.พ.','มี.ค.','เม.ย.','พ.ค.','มิ.ย.','ก.ค.','ส.ค.','ก.ย.','ต.ค.','พ.ย.','ธ.ค.']
        return f"{int(d)} {mn[int(m)]} {int(y)+543}"
    except: return s

def norm(s):
    if not s: return s
    s = re.sub(r'า้', 'า้', s)
    fixes = [
        (r'ตน้ ', 'ต้น '),
        (r'รา้น', 'ร้าน'), (r'บา้น', 'บ้าน'), (r'นา้', 'น้ำ'),
        (r'ทา่ ', 'ท่า'), (r'คา้', 'ค้า'), (r'ผา้', 'ผ้า'),
        (r'หนา้', 'หน้า'), (r'ลา้', 'ล้า'), (r'ขา้', 'ข้า'),
        (r'ขา่ ', 'ข่า'), (r'รา่ ', 'ร่า'),
        (r'พรพมิ ล', 'พรพิมล'),
        (r'วาสนา จางนะ', 'วาสนา จางนะ'),
        (r'จอมแจง้', 'จอมแจ้ง'),
        (r'บารเ์บอร์?', 'บาร์เบอร์'),
        (r'พศิ ษิ ฐ์?', 'พิษฐ์'),
        (r'รองเทา้', 'รองเท้า'),
        (r'ถกู ', 'ถูก'),
        (r'มดแดง', 'มดแดง'),
        (r'เมอื ง', 'เมือง'),
        (r'แมส่ ะเรยี ง', 'แม่สะเรียง'),
        (r'แมฮ่ อ่ งสอน', 'แม่ฮ่องสอน'),
        (r'วหิ ารแดง', 'วิหารแดง'),
        (r'สวุ รรณศร', 'สุวรรณศร'),
        (r'บรษิ ทั ', 'บริษัท '), (r'บรษิัท', 'บริษัท'),
        (r'จาํ กดั', 'จำกัด'), (r'จํากดั', 'จำกัด'),
        (r'คณุ ', 'คุณ '), (r'คณุ$', 'คุณ'),
        (r'ชลบรุ ี', 'ชลบุรี'), (r'ชลบรีุ', 'ชลบุรี'),
        (r'สระบรุ ี', 'สระบุรี'),
        (r'เพชรบรุ ี', 'เพชรบุรี'),
        (r'สพุ รรณบรุ ี', 'สุพรรณบุรี'),
        (r'ประจวบครีขี ันธ์', 'ประจวบคีรีขันธ์'),
        (r'กําแพงเพชร', 'กำแพงเพชร'),
        (r'นครสวรรค์', 'นครสวรรค์'),
        (r'เชยี งราย', 'เชียงราย'), (r'เชยงราย ี', 'เชียงราย'),
        (r'ภเูก็ต', 'ภูเก็ต'),
        (r'ระยอง', 'ระยอง'),
        (r'ทา่ ยาง', 'ท่ายาง'),
        (r'สบเมย', 'สบเมย'),
        (r'บา้นสวน', 'บ้านสวน'),
        (r'ปราณบรุ ี', 'ปราณบุรี'),
        (r'โพนพสิ ัย', 'โพนพิสัย'),
        (r'อู่ทอง', 'อู่ทอง'),
        (r'ขาณวุรลักษบรุ ี', 'ขาณุวรลักษบุรี'),
        (r'ซปุ เปอร์', 'ซุปเปอร์'),
        (r'ไทยนยิ ม', 'ไทยนิยม'),
        (r'กําไลบตู คิ', 'กำไลบูติก'),
        (r'สริิภพ', 'สิริภพ'),
        (r'แคลว้', 'แคล้ว'),
        (r'อยีมเจรญิ', 'อ้ยีมเจริญ'),
        (r'ปัญชรัสมิ', 'ปัญชรัสมิ์'),
        (r'\s+', ' '),
    ]
    for p, r in fixes:
        s = re.sub(p, r, s)
    return s.strip()


def detect_type(barcode):
    p = barcode[:3]
    if barcode[:2] == '11': return 'canvas'
    if p in ('122','123'): return 'foam212'
    if p in ('120','121'): return 'foam200'
    return 'gift'

def get_subtype(barcode, desc):
    if '205S' in desc: return '205S'
    if '205R' in desc: return '205R'
    if barcode[:3] == '123' or '213' in desc: return '213'
    if barcode[:3] == '122' or '212' in desc: return '212'
    if '200' in desc: return '200'
    return desc.split()[0] if desc else 'อื่นๆ'

def extract_header_lines(page):
    words = page.extract_words(x_tolerance=3, y_tolerance=3)
    rows = {}
    for w in words:
        y = round(w['top']/5)*5
        rows.setdefault(y,[]).append(w['text'])
    lines = []
    for y in sorted(rows):
        raw = ' '.join(rows[y])
        raw = re.sub(r'\(cid:\d+\)', '', raw)
        l = re.sub(r'\s+', ' ', raw).strip()
        if l: lines.append(l)
    return lines

def parse_pdf(file_bytes):
    page_data = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            hlines = extract_header_lines(page)
            txt = page.extract_text() or ''
            txt_c = re.sub(r'\(cid:\d+\)', '', txt)
            txt_c = re.sub(r'\s+', ' ', txt_c)
            page_data.append((hlines, txt_c))

    if not page_data: return []

    def get_ifo(lines):
        for l in lines[:10]:
            m = re.search(r'IFO-\d+', l)
            if m: return m.group()
        return None

    groups = {}; order = []
    for hlines, txt in page_data:
        ifo_id = get_ifo(hlines)
        if not ifo_id: continue
        if ifo_id not in groups:
            groups[ifo_id] = {'hlines':[], 'text':''}
            order.append(ifo_id)
        groups[ifo_id]['hlines'].extend(hlines)
        groups[ifo_id]['text'] += txt + ' '

    if not groups: return []

    docs = []
    for ifo_id in order:
        g = groups[ifo_id]
        doc = {'docId':ifo_id,'date':'','customer':'','province':'','amphoe':'','items':[]}
        lines = g['hlines']
        full_text = g['text']

        pat = r'(\d{1,2})\s+(' + '|'.join(re.escape(k) for k in MONTH_MAP) + r')\s+(\d{4})'
        for line in lines[:15]:
            m = re.search(pat, line)
            if m:
                d,mo,y = m.group(1),m.group(2),int(m.group(3))
                if y > 2500: y -= 543
                doc['date'] = f"{y}-{MONTH_MAP[mo]}-{d.zfill(2)}"
                break

        for line in lines[:20]:
            m = re.search(r'ช\S*\s+ล\S*\s+ค\S*\s*:\s*(.+?)(?:\s+ว\S*\s+ท|$)', line)
            if m:
                doc['customer'] = norm(m.group(1).strip())
                break

        for line in lines[:15]:
            mp = re.search(r'จ\.([ก-๙]+(?:\s+[ก-๙]+)?)', line)
            ma = re.search(r'อ\.([ก-๙]+(?:\s+[ก-๙]+)?)(?:\s+จ\.)?', line)
            if mp: doc['province'] = norm(re.sub(r'\d+.*','',mp.group(1)).strip())
            if ma: doc['amphoe'] = norm(re.sub(r'\s*จ$','',re.sub(r'\d+.*','',ma.group(1)).strip()).strip())
            if doc['province']: break

        seen = set()

        gift_pat = re.compile(r'(\d{1,4})\s+คู่\s+0\.00\s+0\.00\s+ของแถม\s+\d+\s+((1[12]\d{7}))\s+(.+?)(?=\d+\s+คู่|Z0001|$)')
        gift_items = set()
        for m in gift_pat.finditer(full_text):
            qty, bc, desc_raw = int(m.group(1)), m.group(2), m.group(4).strip()
            if qty <= 0 or qty > 9999: continue
            key = (bc, qty)
            if key in seen: continue
            seen.add(key); gift_items.add(key)
            desc = re.sub(r'\s+\d+(\.\d+)?(\s+\d+(\.\d+)?)*$','',desc_raw).strip()
            desc = re.sub(r'^\d+\s+','',desc).strip()
            ptype = detect_type(bc)
            subtype = get_subtype(bc, desc)
            doc['items'].append({'desc':desc,'type':ptype,'subtype':subtype,'qty':qty,'gift':True})

        norm_pat = re.compile(r'(?<![\d])((1[12]\d{7}))\s+(.+?)\s+(\d{1,4})\s+คู่(.{0,80}?)(?=(?:1[12]\d{7})|Z0001|$)')
        for m in norm_pat.finditer(full_text):
            bc, desc_raw, qty = m.group(1), m.group(3).strip(), int(m.group(4))
            after = m.group(5) or ''
            if qty <= 0 or qty > 9999: continue
            key = (bc, qty)
            if key in seen: continue
            seen.add(key)
            desc = re.sub(r'\s+\d+(\.\d+)?(\s+\d+(\.\d+)?)*$','',desc_raw).strip()
            desc = re.sub(r'^\d+\s+','',desc).strip()
            ptype = detect_type(bc)
            subtype = get_subtype(bc, desc)
            is_gift = bool(re.search(r'ของแถม', after))
            if not is_gift:
                prices = re.findall(r'(\d+\.\d+)', after)
                if len(prices) >= 2 and float(prices[0]) == 0.0 and float(prices[1]) == 0.0:
                    is_gift = True
            doc['items'].append({'desc':desc,'type':ptype,'subtype':subtype,'qty':qty,'gift':is_gift})

        if doc['items']:
            docs.append(doc)
    return docs


def calc_canvas(n): return {'qty':n,'boxes':n//BOX_CANVAS,'rem':n%BOX_CANVAS}
def calc_foam200(n):
    doz=n//12; sacks=doz//10; rem_doz=doz%10
    return {'qty':n,'doz':doz,'sacks':sacks,'rem_doz':rem_doz}
def calc_foam212(n): return {'qty':n,'boxes':n//BOX_212,'rem_doz':(n%BOX_212)//12}

def agg(doc):
    doc['_ct']      = sum(x['qty'] for x in doc['items'] if x['type']=='canvas' and not x['gift'])
    doc['_ct_gift'] = sum(x['qty'] for x in doc['items'] if x['type']=='canvas' and x['gift'])
    doc['_ft2']     = sum(x['qty'] for x in doc['items'] if x['type']=='foam200' and not x['gift'])
    doc['_ft2_gift']= sum(x['qty'] for x in doc['items'] if x['type']=='foam200' and x['gift'])
    doc['_ft3_212'] = sum(x['qty'] for x in doc['items'] if x['type']=='foam212' and x['subtype']=='212' and not x['gift'])
    doc['_ft3_213'] = sum(x['qty'] for x in doc['items'] if x['type']=='foam212' and x['subtype']=='213' and not x['gift'])
    doc['_ft3_212_gift'] = sum(x['qty'] for x in doc['items'] if x['type']=='foam212' and x['subtype']=='212' and x['gift'])
    doc['_ft3_213_gift'] = sum(x['qty'] for x in doc['items'] if x['type']=='foam212' and x['subtype']=='213' and x['gift'])

def build_excel(docs):
    wb = Workbook()

    # ── Colors & Styles ──
    BLUE='1B4F8A'; WHITE='FFFFFF'
    CG='E2EFDA'; CGH='375623'
    CF='DBEAFE'; CFH='1E3A8A'
    C212='D0E4FF'; C212H='1E40AF'
    C213='EDE9FE'; C213H='5B21B6'
    CGIFT='FFF9C4'; CALT='F5F8FA'

    thin=Side(style='thin',color='BBBBBB')
    bdr=Border(left=thin,right=thin,top=thin,bottom=thin)

    FONT = 'TH Sarabun New'
    def hf(sz=11,color=WHITE,bold=True):
        return Font(name=FONT,bold=bold,color=color,size=sz)
    def nf(sz=11,bold=False,color='1E293B'):
        return Font(name=FONT,size=sz,bold=bold,color=color)
    def fl(c): return PatternFill('solid',start_color=c,end_color=c)
    def al(h='center',wrap=True):
        return Alignment(horizontal=h,vertical='center',wrap_text=wrap)

    def wcell(ws,r,c,val,font=None,fill=None,align=None,border=None):
        cell = ws.cell(row=r,column=c,value=val)
        if font:  cell.font=font
        if fill:  cell.fill=fill
        if align: cell.alignment=align
        if border:cell.border=border
        return cell

    def mwrite(ws,r1,c1,r2,c2,val,font=None,fill=None,align=None):
        if r1==r2 and c1==c2:
            ws.cell(row=r1,column=c1,value=val)
        else:
            ws.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)
            ws.cell(row=r1,column=c1,value=val)
        for rr in range(r1,r2+1):
            for cc in range(c1,c2+1):
                cell=ws.cell(row=rr,column=cc)
                if font:  cell.font=font
                if fill:  cell.fill=fill
                if align: cell.alignment=align
                cell.border=bdr

    # ── หา subtypes ──
    canvas_subs = sorted(set(x['subtype'] for d in docs for x in d['items'] if x['type']=='canvas'))
    has212 = any(x['subtype']=='212' for d in docs for x in d['items'] if x['type']=='foam212')
    has213 = any(x['subtype']=='213' for d in docs for x in d['items'] if x['type']=='foam212')

    # ══════════════════════════════════════════
    # SHEET 1: สรุปรายเอกสาร
    # ══════════════════════════════════════════
    ws1 = wb.active; ws1.title='สรุปรายเอกสาร'

    grp_cols = []
    grp_cols.append({'sc':1,'ec':5,'label':'ข้อมูลเอกสาร','hc':BLUE,'dc':BLUE,
                     'subs':['วันที่','เลขที่IFO','ชื่อลูกค้า','อำเภอ','จังหวัด']})
    col=6
    for sub in canvas_subs:
        grp_cols.append({'sc':col,'ec':col+5,'label':f'ผ้าใบ {sub} (12 คู่/กล่อง)','hc':CGH,'dc':CG,
                         'subs':['คู่','กล่อง','เศษ','ของแถม คู่','กล่อง','เศษ'],
                         'key':f'canvas_{sub}','gift_key':f'canvas_{sub}_gift','type':'canvas'})
        col+=6
    grp_cols.append({'sc':col,'ec':col+6,'label':'ฟองน้ำ 200 (120 คู่/กระสอบ)','hc':CFH,'dc':CF,
                     'subs':['คู่','โหล','กระสอบ','เศษโหล','ของแถม คู่','กระสอบ','เศษโหล'],
                     'key':'foam200','gift_key':'foam200_gift','type':'foam200'})
    col+=7
    if has212:
        grp_cols.append({'sc':col,'ec':col+4,'label':'ฟองน้ำ 212 (24 คู่/กล่อง)','hc':C212H,'dc':C212,
                         'subs':['คู่','กล่อง','ของแถม คู่','กล่อง','เศษ'],
                         'key':'foam212','gift_key':'foam212_gift','type':'foam212'})
        col+=5
    if has213:
        grp_cols.append({'sc':col,'ec':col+4,'label':'ฟองน้ำ 213 (24 คู่/กล่อง)','hc':C213H,'dc':C213,
                         'subs':['คู่','กล่อง','ของแถม คู่','กล่อง','เศษ'],
                         'key':'foam213','gift_key':'foam213_gift','type':'foam213'})
        col+=5
    grp_cols.append({'sc':col,'ec':col,'label':'รวม','hc':BLUE,'dc':BLUE,'subs':['รวมคู่'],
                     'key':'total','type':'sum'})
    NCOLS=col

    mwrite(ws1,1,1,1,NCOLS,'สรุปโหลดสินค้า — บริษัท นันยางมาร์เก็ตติ้ง จำกัด',
           font=hf(14),fill=fl(BLUE),align=al())
    ws1.row_dimensions[1].height=28
    mwrite(ws1,2,1,2,NCOLS,f'พิมพ์: {datetime.now().strftime("%d/%m/%Y %H:%M")}',
           font=nf(10),align=al('right'))
    ws1.row_dimensions[2].height=16
    for g in grp_cols:
        mwrite(ws1,3,g['sc'],3,g['ec'],g['label'],font=hf(11),fill=fl(g['hc']),align=al())
    ws1.row_dimensions[3].height=22
    col_idx=1
    col_dc_map={}
    for g in grp_cols:
        for si,sub in enumerate(g['subs']):
            wcell(ws1,4,col_idx,sub,font=hf(10),fill=fl(g['hc']),align=al(),border=bdr)
            col_dc_map[col_idx]=g['dc']
            col_idx+=1
    ws1.row_dimensions[4].height=34

    # ── Data rows ──
    tot=defaultdict(int)
    for i,doc in enumerate(docs):
        r=5+i
        bg='FFFFFF' if i%2==0 else CALT

        # ── คำนวณ canvas/foam แยก gift (define ก่อนใช้ทุกกรณี) ──
        canvas_qty=defaultdict(int); canvas_gift=defaultdict(int)
        for x in doc['items']:
            if x['type']=='canvas':
                if x['gift']: canvas_gift[x['subtype']]+=x['qty']
                else: canvas_qty[x['subtype']]+=x['qty']
        ft2=doc['_ft2']; ft2g=doc.get('_ft2_gift',0)
        f212=doc['_ft3_212']; f212g=doc.get('_ft3_212_gift',0)
        f213=doc['_ft3_213']; f213g=doc.get('_ft3_213_gift',0)
        grand=(doc['_ct']+doc.get('_ct_gift',0)+ft2+ft2g+f212+f212g+f213+f213g)

        # ── สะสม totals ──
        for sub in canvas_subs:
            tot[f'c_{sub}']+=canvas_qty.get(sub,0)
            tot[f'cg_{sub}']+=canvas_gift.get(sub,0)
        tot['ft2']+=ft2; tot['ft2g']+=ft2g
        tot['f212']+=f212; tot['f212g']+=f212g
        tot['f213']+=f213; tot['f213g']+=f213g

        row_vals=[th_date(doc['date']),doc['docId'],doc['customer'],
                  doc.get('amphoe',''),doc.get('province','')]
        row_bgs=[bg]*5

        for sub in canvas_subs:
            ct=canvas_qty.get(sub,0); cg=canvas_gift.get(sub,0)
            cc=calc_canvas(ct); gc=calc_canvas(cg)
            row_vals+=[ct or '-',cc['boxes'] or '-',cc['rem'] or '-',
                       cg or '-',gc['boxes'] or '-',gc['rem'] or '-']
            row_bgs+=[CG if ct else bg]*3+[CGIFT if cg else bg]*3

        cf2=calc_foam200(ft2); gf2=calc_foam200(ft2g)
        row_vals+=[ft2 or '-',cf2['doz'] or '-',cf2['sacks'] or '-',cf2['rem_doz'] or '-',
                   ft2g or '-',gf2['sacks'] or '-',gf2['rem_doz'] or '-']
        row_bgs+=[CF if ft2 else bg]*4+[CGIFT if ft2g else bg]*3

        if has212:
            c2=calc_foam212(f212); g2=calc_foam212(f212g)
            row_vals+=[f212 or '-',c2['boxes'] or '-',f212g or '-',g2['boxes'] or '-',c2['rem_doz'] or '-']
            row_bgs+=[C212 if f212 else bg]*2+[CGIFT if f212g else bg]*2+[bg]
        if has213:
            c3=calc_foam212(f213); g3=calc_foam212(f213g)
            row_vals+=[f213 or '-',c3['boxes'] or '-',f213g or '-',g3['boxes'] or '-',c3['rem_doz'] or '-']
            row_bgs+=[C213 if f213 else bg]*2+[CGIFT if f213g else bg]*2+[bg]

        row_vals+=[grand]; row_bgs+=['D6E4F0']

        for ci,(val,bgc) in enumerate(zip(row_vals,row_bgs),1):
            wcell(ws1,r,ci,val,
                  font=nf(11,bold=(ci==NCOLS)),
                  fill=fl(bgc),
                  align=al('left' if ci<=3 else 'center'),
                  border=bdr)
        ws1.row_dimensions[r].height=22

    # Summary row
    sr=5+len(docs)
    mwrite(ws1,sr,1,sr,5,'รวมทั้งหมด',font=hf(11),fill=fl(BLUE),align=al())
    ci=6
    for sub in canvas_subs:
        cc=calc_canvas(tot.get(f'c_{sub}',0)); gc=calc_canvas(tot.get(f'cg_{sub}',0))
        ct=tot.get(f'c_{sub}',0); cg=tot.get(f'cg_{sub}',0)
        for v in [ct or '-',cc['boxes'] or '-',cc['rem'] or '-',
                  cg or '-',gc['boxes'] or '-',gc['rem'] or '-']:
            wcell(ws1,sr,ci,v,font=hf(11),fill=fl(BLUE),align=al('center'),border=bdr); ci+=1
    cf2=calc_foam200(tot.get('ft2',0)); gf2=calc_foam200(tot.get('ft2g',0))
    for v in [tot.get('ft2',0) or '-',cf2['doz'] or '-',cf2['sacks'] or '-',cf2['rem_doz'] or '-',
              tot.get('ft2g',0) or '-',gf2['sacks'] or '-',gf2['rem_doz'] or '-']:
        wcell(ws1,sr,ci,v,font=hf(11),fill=fl(BLUE),align=al('center'),border=bdr); ci+=1
    if has212:
        c2=calc_foam212(tot.get('f212',0)); g2=calc_foam212(tot.get('f212g',0))
        for v in [tot.get('f212',0) or '-',c2['boxes'] or '-',
                  tot.get('f212g',0) or '-',g2['boxes'] or '-',c2['rem_doz'] or '-']:
            wcell(ws1,sr,ci,v,font=hf(11),fill=fl(BLUE),align=al('center'),border=bdr); ci+=1
    if has213:
        c3=calc_foam212(tot.get('f213',0)); g3=calc_foam212(tot.get('f213g',0))
        for v in [tot.get('f213',0) or '-',c3['boxes'] or '-',
                  tot.get('f213g',0) or '-',g3['boxes'] or '-',c3['rem_doz'] or '-']:
            wcell(ws1,sr,ci,v,font=hf(11),fill=fl(BLUE),align=al('center'),border=bdr); ci+=1
    grand_tot = sum(tot.get(k,0) for k in ['ft2','ft2g','f212','f212g','f213','f213g'])
    grand_tot += sum(tot.get(f'c_{s}',0)+tot.get(f'cg_{s}',0) for s in canvas_subs)
    wcell(ws1,sr,ci,grand_tot,font=hf(11),fill=fl(BLUE),align=al('center'),border=bdr)
    ws1.row_dimensions[sr].height=24

    base_w=[12,14,26,12,14]
    for _ in canvas_subs: base_w+=[8,8,8,8,8,8]
    base_w+=[8,8,10,10,8,10,10]
    if has212: base_w+=[8,8,8,8,8]
    if has213: base_w+=[8,8,8,8,8]
    base_w+=[10]
    for ci,w in enumerate(base_w,1):
        ws1.column_dimensions[get_column_letter(ci)].width=w

    # ══════════════════════════════════════════
    # SHEET 2: สรุปรวม
    # ══════════════════════════════════════════
    ws2=wb.create_sheet('สรุปรวม')

    mwrite(ws2,1,1,1,8,f'สรุปรวมทุกเอกสาร ({len(docs)} IFO) — {datetime.now().strftime("%d/%m/%Y %H:%M")}',
           font=hf(13),fill=fl(BLUE),align=al())
    ws2.row_dimensions[1].height=26

    blocks=[]
    for sub in canvas_subs:
        ct=tot[f'c_{sub}']; cc=calc_canvas(ct)
        blocks.append((f'ผ้าใบ {sub} (ปกติ)',f'{ct} คู่ / {cc["boxes"]} กล่อง',CGH,CG))
    if tot['ft2']:
        cf2=calc_foam200(tot['ft2'])
        txt=f'{tot["ft2"]} คู่ / {cf2["sacks"]} กระสอบ'
        if cf2['rem_doz']: txt+=f' เศษ {cf2["rem_doz"]} โหล'
        blocks.append(('ฟองน้ำ 200 (ปกติ)',txt,CFH,CF))
    if tot['f212']:
        c2=calc_foam212(tot['f212'])
        blocks.append(('ฟองน้ำ 212 (ปกติ)',f'{tot["f212"]} คู่ / {c2["boxes"]} กล่อง',C212H,C212))
    if tot['f213']:
        c3=calc_foam212(tot['f213'])
        blocks.append(('ฟองน้ำ 213 (ปกติ)',f'{tot["f213"]} คู่ / {c3["boxes"]} กล่อง',C213H,C213))

    col=1
    for label,val,hc,dc in blocks:
        mwrite(ws2,3,col,3,col+1,label,font=hf(11),fill=fl(hc),align=al())
        mwrite(ws2,4,col,4,col+1,val,font=nf(12,bold=True),fill=fl(dc),align=al())
        ws2.column_dimensions[get_column_letter(col)].width=16
        ws2.column_dimensions[get_column_letter(col+1)].width=16
        col+=2
    ws2.row_dimensions[3].height=22; ws2.row_dimensions[4].height=26

    r=6
    mwrite(ws2,r,1,r,8,'รายละเอียดตามชนิดสินค้า (รวมทุก IFO)',font=hf(12),fill=fl(BLUE),align=al())
    ws2.row_dimensions[r].height=22; r+=1

    det_hdrs=['ชนิดสินค้า','ประเภท','รวม คู่','โหล','กล่อง/กระสอบ','เศษโหล','เศษคู่','หน่วย']
    for ci,h in enumerate(det_hdrs,1):
        wcell(ws2,r,ci,h,font=hf(11),fill=fl(BLUE),align=al(),border=bdr)
    ws2.row_dimensions[r].height=22; r+=1

    subtype_data=defaultdict(lambda:defaultdict(int))
    for doc in docs:
        for x in doc['items']:
            subtype_data[(x['subtype'],x['type'],x['gift'])]['qty']+=x['qty']

    sorted_keys=sorted(subtype_data.keys(),key=lambda x:({'canvas':0,'foam200':1,'foam212':2}.get(x[1],3),int(x[2]),x[0]))
    last_gift=None
    for key in sorted_keys:
        sub,ptype,gift=key
        qty=subtype_data[key]['qty']
        if qty==0: continue
        if gift and last_gift==False:
            ws2.row_dimensions[r].height=6; r+=1
        last_gift=gift
        doz=qty//12
        if ptype=='canvas':
            cc=calc_canvas(qty); pack=cc['boxes']; rem_doz='-'; rem_pair=cc['rem'] or '-'; unit='กล่อง'; dc=CGIFT if gift else CG
        elif ptype=='foam200':
            cf=calc_foam200(qty); pack=cf['sacks']; rem_doz=cf['rem_doz'] or '-'; rem_pair='-'; unit='กระสอบ'; dc=CGIFT if gift else CF
        else:
            cf=calc_foam212(qty); pack=cf['boxes']; rem_doz=cf['rem_doz'] or '-'; rem_pair='-'; unit='กล่อง'
            dc=(CGIFT if gift else (C212 if sub=='212' else C213))
        row_v=[sub,'ของแถม' if gift else 'ปกติ',qty,doz,pack,rem_doz,rem_pair,unit]
        for ci,v in enumerate(row_v,1):
            wcell(ws2,r,ci,v,font=nf(11,bold=(ci==1)),fill=fl(dc),align=al('left' if ci<=2 else 'center'),border=bdr)
        ws2.row_dimensions[r].height=20; r+=1

    for ci,w in enumerate([16,10,10,8,14,10,10,10],1):
        ws2.column_dimensions[get_column_letter(ci)].width=w

    # ══════════════════════════════════════════
    # SHEET 3: รายละเอียด
    # ══════════════════════════════════════════
    ws3=wb.create_sheet('รายละเอียด')
    mwrite(ws3,1,1,1,8,'รายละเอียดสินค้าทุกรายการ',font=hf(13),fill=fl(BLUE),align=al())
    ws3.row_dimensions[1].height=24

    dh=['วันที่','เลขที่IFO','ชื่อลูกค้า','รายการสินค้า','ประเภท','คู่','โหล','กล่อง/กระสอบ/กล่อง']
    for ci,h in enumerate(dh,1):
        wcell(ws3,2,ci,h,font=hf(11),fill=fl(BLUE),align=al(),border=bdr)
    ws3.row_dimensions[2].height=22

    TYPE_LABEL={'canvas':'ผ้าใบ','foam200':'ฟองน้ำ 200','foam212':'ฟองน้ำ 212/213'}
    TYPE_COLOR={'canvas':CG,'foam200':CF,'foam212':C212}

    r3=3
    for doc in docs:
        for x in doc['items']:
            qty=x['qty']; doz=qty//12; ptype=x['type']
            if ptype=='canvas':
                cc=calc_canvas(qty); load=f'{cc["boxes"]} กล่อง'+(f' เศษ {cc["rem"]} คู่' if cc['rem'] else '')
            elif ptype=='foam200':
                cf=calc_foam200(qty); load=f'{cf["sacks"]} กระสอบ'+(f' เศษ {cf["rem_doz"]} โหล' if cf['rem_doz'] else '')
            else:
                cf=calc_foam212(qty); load=f'{cf["boxes"]} กล่อง'+(f' เศษ {cf["rem_doz"]} โหล' if cf['rem_doz'] else '')
            label=TYPE_LABEL.get(ptype,'อื่นๆ')
            dc=CGIFT if x['gift'] else (C213 if x['subtype']=='213' else TYPE_COLOR.get(ptype,'FFFFFF'))
            row_v=[th_date(doc['date']),doc['docId'],doc['customer'],x['desc'],label,qty,doz,load]
            for ci,v in enumerate(row_v,1):
                wcell(ws3,r3,ci,v,font=nf(10),fill=fl(dc),align=al('left' if ci<=4 else 'center'),border=bdr)
            ws3.row_dimensions[r3].height=18; r3+=1

    for ci,w in enumerate([12,14,26,36,14,8,8,20],1):
        ws3.column_dimensions[get_column_letter(ci)].width=w

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


# ── UI ──
st.markdown("""
<div style="background:linear-gradient(135deg,#1B4F8A,#163D6E);color:white;border-radius:12px;padding:16px 24px;margin-bottom:20px">
<h1 style="font-size:20px;font-weight:700;margin:0">📦 คำนวณโหลดสินค้า IFO</h1>
<p style="font-size:13px;opacity:.7;margin:4px 0 0">อัปโหลด PDF → คำนวณอัตโนมัติ → ดาวน์โหลด Excel</p>
</div>""", unsafe_allow_html=True)

if 'ukey' not in st.session_state: st.session_state.ukey=0

c1,c2=st.columns([4,1])
with c1:
    uploaded=st.file_uploader("📂 ลาก PDF มาวาง (รองรับ PDF รวมหลาย IFO)",
        type=['pdf'],accept_multiple_files=True,key=f"up_{st.session_state.ukey}")
with c2:
    st.markdown('<br>',unsafe_allow_html=True)
    if st.button("🗑️ ล้างข้อมูล",use_container_width=True):
        st.session_state.ukey+=1; st.rerun()

if uploaded:
    docs=[]; errors=[]
    with st.spinner('🔍 กำลังอ่าน PDF...'):
        for f in uploaded:
            try:
                results=parse_pdf(f.read())
                for doc in results:
                    agg(doc); doc['_file']=f.name; docs.append(doc)
            except Exception as e:
                errors.append(f"{f.name}: {e}")

    for e in errors: st.error(f"❌ {e}")

    if docs:
        docs.sort(key=lambda x:x['docId'])
        tot_c=sum(d['_ct'] for d in docs)
        tot_f=sum(d['_ft2'] for d in docs)
        tot_212=sum(d['_ft3_212'] for d in docs)
        tot_213=sum(d['_ft3_213'] for d in docs)

        canvas_subs_all=defaultdict(int)
        for d in docs:
            for x in d['items']:
                if x['type']=='canvas': canvas_subs_all[x['subtype']]+=x['qty']

        metric_items=[("📄 เอกสาร",f"{len(docs)} IFO")]
        for sub,qty in sorted(canvas_subs_all.items()):
            metric_items.append((f"🟢 ผ้าใบ {sub}",f"{qty} คู่"))
        metric_items.append(("🔵 ฟองน้ำ 200",f"{tot_f} คู่"))
        if tot_212: metric_items.append(("🔵 ฟองน้ำ 212",f"{tot_212} คู่"))
        if tot_213: metric_items.append(("🔵 ฟองน้ำ 213",f"{tot_213} คู่"))

        cols=st.columns(len(metric_items))
        for i,(label,val) in enumerate(metric_items):
            cols[i].metric(label,val)

        st.divider()

        for doc in docs:
            with st.expander(f"📄 {doc['docId']} — {doc['customer']} ({th_date(doc['date'])})",expanded=True):
                c1,c2,c3=st.columns(3)
                c1.write(f"**ลูกค้า:** {doc['customer']}")
                c2.write(f"**อ.{doc.get('amphoe','')} จ.{doc.get('province','')}**")
                c3.write(f"**รวม:** {doc['_ct']+doc['_ft2']+doc['_ft3_212']+doc['_ft3_213']} คู่")
                if doc['_ct']:
                    cc=calc_canvas(doc['_ct'])
                    st.success(f"🟢 **ผ้าใบ:** {doc['_ct']} คู่ / {cc['boxes']} กล่อง" + (f" เศษ {cc['rem']} คู่" if cc['rem'] else ""))
                if doc['_ft2']:
                    cf=calc_foam200(doc['_ft2'])
                    st.info(f"🔵 **ฟองน้ำ 200:** {doc['_ft2']} คู่ / {cf['sacks']} กระสอบ เศษ {cf['rem_doz']} โหล")
                if doc['_ft3_212']:
                    c2=calc_foam212(doc['_ft3_212'])
                    st.info(f"🔵 **ฟองน้ำ 212:** {doc['_ft3_212']} คู่ / {c2['boxes']} กล่อง")
                if doc['_ft3_213']:
                    c3=calc_foam212(doc['_ft3_213'])
                    st.info(f"🔵 **ฟองน้ำ 213:** {doc['_ft3_213']} คู่ / {c3['boxes']} กล่อง")
                gift_total = doc.get('_ct_gift',0)+doc.get('_ft2_gift',0)+doc.get('_ft3_212_gift',0)+doc.get('_ft3_213_gift',0)
                if gift_total:
                    parts = []
                    if doc.get('_ct_gift'):
                        gc=calc_canvas(doc['_ct_gift'])
                        parts.append(f"ผ้าใบ {doc['_ct_gift']} คู่ / {gc['boxes']} กล่อง")
                    if doc.get('_ft2_gift'):
                        gf=calc_foam200(doc['_ft2_gift'])
                        parts.append(f"ฟองน้ำ200 {doc['_ft2_gift']} คู่ / {gf['sacks']} กระสอบ")
                    if doc.get('_ft3_212_gift'):
                        g2=calc_foam212(doc['_ft3_212_gift'])
                        parts.append(f"ฟองน้ำ212 {doc['_ft3_212_gift']} คู่ / {g2['boxes']} กล่อง")
                    if doc.get('_ft3_213_gift'):
                        g3=calc_foam212(doc['_ft3_213_gift'])
                        parts.append(f"ฟองน้ำ213 {doc['_ft3_213_gift']} คู่ / {g3['boxes']} กล่อง")
                    st.warning(f"🎁 **ของแถม:** {' | '.join(parts)}")

        excel=build_excel(docs)
        fname=f"IFO_โหลดสินค้า_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        st.download_button("📥 ดาวน์โหลด Excel",data=excel,file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,type="primary")
    else:
        if not errors: st.warning("⚠️ ไม่พบข้อมูล IFO ในไฟล์ที่อัปโหลด")
